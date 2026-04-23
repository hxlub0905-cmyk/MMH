"""Export measurement results to a comprehensive multi-sheet Excel workbook.

Sheet layout
────────────
  "All Measurements"  — every MeasurementRecord with CD line positions,
                        MIN rows highlighted orange, MAX rows blue.
  "Image Summary"     — one row per image: count, mean, std, min/max CD values
                        and their XY positions (relative to image top-left).
                        MIN-CD and MAX-CD cells are highlighted for easy search.
  "Statistics"        — overall stats per recipe (or single block).
"""

from __future__ import annotations
from pathlib import Path


def _require_pandas():
    try:
        import pandas as pd
        return pd
    except ImportError:
        raise ImportError(
            "pandas 未安裝或無法載入。\n"
            "請在正確的虛擬環境中執行：\n"
            "  pip install \"pandas>=2.0\" openpyxl\n"
            "若使用 PyCharm，請確認 Project Interpreter 已設定正確。"
        )


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError(
            "openpyxl 未安裝或無法載入。\n"
            "請在正確的虛擬環境中執行：\n"
            "  pip install \"pandas>=2.0\" openpyxl"
        )


# ── colour constants ──────────────────────────────────────────────────────────
_MIN_FILL   = "FFD580"   # warm orange  (MIN CD)
_MAX_FILL   = "AED6F1"   # sky blue     (MAX CD)
_HEADER_FILL = "F7E0C8"  # pale warm    (column headers)
_SUMMARY_HEADER_FILL = "D5E8D4"  # pale green  (Image Summary headers)
_STAT_HEADER_FILL    = "DAE8FC"  # pale blue   (Statistics headers)


# ── legacy (old-style results list) entry point ───────────────────────────────

def export_excel(results: list[dict], out_path: Path, nm_per_pixel: float) -> None:
    pd = _require_pandas()
    _require_openpyxl()
    from ._common import results_to_dataframe
    df = results_to_dataframe(results, nm_per_pixel)
    ok = df[df["status"] == "OK"]["y_cd_nm"].dropna()

    stats = {
        "Metric": ["Count", "Mean (nm)", "Median (nm)", "Q25 (nm)", "Q75 (nm)",
                   "Std Dev (nm)", "3-Sigma (nm)", "Min (nm)", "Max (nm)"],
        "Value": [
            len(ok),
            round(ok.mean(), 3) if len(ok) else "N/A",
            round(ok.median(), 3) if len(ok) else "N/A",
            round(ok.quantile(0.25), 3) if len(ok) else "N/A",
            round(ok.quantile(0.75), 3) if len(ok) else "N/A",
            round(ok.std(), 3) if len(ok) else "N/A",
            round(ok.std() * 3, 3) if len(ok) else "N/A",
            round(ok.min(), 3) if len(ok) else "N/A",
            round(ok.max(), 3) if len(ok) else "N/A",
        ],
    }
    df_stats = pd.DataFrame(stats)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Measurements", index=False)
        df_stats.to_excel(writer, sheet_name="Statistics", index=False)

        wb = writer.book
        ws = wb["Measurements"]
        from openpyxl.styles import PatternFill
        red_fill = PatternFill("solid", fgColor="FFAAAA")
        blue_fill = PatternFill("solid", fgColor="AAAAFF")

        for row_idx, flag in enumerate(df["flag"], start=2):
            if flag == "MIN":
                for cell in ws[row_idx]:
                    cell.fill = red_fill
            elif flag == "MAX":
                for cell in ws[row_idx]:
                    cell.fill = blue_fill


# ── new-style (MeasurementRecord list) entry point ────────────────────────────

def export_excel_from_records(
    records: list,
    out_path: Path,
    image_records: list | None = None,
    dataset_label: str = "",
    datasets: list[dict] | None = None,
    meas_mode: str = "all",
) -> None:
    """Create a comprehensive multi-sheet Excel workbook from MeasurementRecord list.

    For multi-dataset runs pass ``datasets`` — a list of dicts each containing:
      ``records`` (list[MeasurementRecord]),
      ``image_records`` (list[ImageRecord]),
      ``dataset_label`` (str).
    When ``datasets`` is provided, ``records``, ``image_records``, and
    ``dataset_label`` are ignored.

    Sheets
    ------
    All Measurements  — full data with CD line positions; MIN/MAX colour-coded
    Image Summary     — per-image stats with MIN/MAX locations (great for search)
    Statistics        — aggregate stats per recipe
    """
    pd = _require_pandas()
    _require_openpyxl()
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from ._common import records_to_dataframe

    if datasets:
        frames = [
            records_to_dataframe(
                ds["records"],
                ds.get("image_records"),
                dataset_label=ds.get("dataset_label", ""),
            )
            for ds in datasets
            if ds.get("records")
        ]
        df = pd.concat(frames, ignore_index=True) if frames else records_to_dataframe([], None)
    else:
        df = records_to_dataframe(records, image_records, dataset_label=dataset_label)

    # ── build Image Summary DataFrame ─────────────────────────────────────────
    summary_rows = _build_image_summary(df)
    df_summary = pd.DataFrame(summary_rows)

    # ── build Statistics DataFrame ─────────────────────────────────────────────
    stats_rows = _build_statistics(df)
    df_stats = pd.DataFrame(stats_rows)

    # ── columns for "All Measurements" sheet ──────────────────────────────────
    meas_cols = [
        "dataset", "image_file", "nm_per_pixel", "recipe_name", "axis",
        "cmg_id", "col_id",
        "cd_px", "cd_nm", "flag",
        "cd_line_x_px", "cd_line_y_px",
        "status",
    ]
    # Only include columns that exist in df (guard against missing cols)
    meas_cols = [c for c in meas_cols if c in df.columns]
    df_meas = df[meas_cols].copy()
    df_meas = _filter_meas_by_mode(df_meas, meas_mode)
    # Drop dataset column if it's all empty (single-dataset runs)
    if "dataset" in df_meas.columns and df_meas["dataset"].astype(str).str.strip().eq("").all():
        df_meas = df_meas.drop(columns=["dataset"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_meas.to_excel(writer, sheet_name="All Measurements", index=False)
        df_summary.to_excel(writer, sheet_name="Image Summary", index=False)
        df_stats.to_excel(writer, sheet_name="Statistics", index=False)

        wb = writer.book

        # ── Style: All Measurements ───────────────────────────────────────────
        ws_meas = wb["All Measurements"]
        _style_header_row(ws_meas, fill_hex=_HEADER_FILL)
        _autofit_columns(ws_meas)

        min_fill = PatternFill("solid", fgColor=_MIN_FILL)
        max_fill = PatternFill("solid", fgColor=_MAX_FILL)

        # Use conditional formatting instead of per-cell coloring:
        # avoids O(rows × cols) cell-style writes that corrupt large files.
        # DifferentialStyle + Rule with bgColor is the correct openpyxl API for
        # formula-based CF; FormulaRule(fill=PatternFill(fgColor=...)) uses an
        # incompatible solid fill that can corrupt workbook state in some versions.
        try:
            from openpyxl.styles.differential import DifferentialStyle
            from openpyxl.formatting.rule import Rule as _Rule
            flag_col_letter = _col_letter(df_meas, "flag")
            if flag_col_letter and 0 < len(df_meas) <= 50_000:
                last_row = len(df_meas) + 1
                last_col = get_column_letter(ws_meas.max_column)
                cf_range = f"A2:{last_col}{last_row}"
                min_dxf = DifferentialStyle(fill=PatternFill(bgColor=_MIN_FILL))
                max_dxf = DifferentialStyle(fill=PatternFill(bgColor=_MAX_FILL))
                ws_meas.conditional_formatting.add(
                    cf_range,
                    _Rule(type="formula", dxf=min_dxf,
                          formula=[f'${flag_col_letter}2="MIN"']),
                )
                ws_meas.conditional_formatting.add(
                    cf_range,
                    _Rule(type="formula", dxf=max_dxf,
                          formula=[f'${flag_col_letter}2="MAX"']),
                )
        except Exception:
            pass  # CF is best-effort; data integrity is not affected

        # Freeze header row
        ws_meas.freeze_panes = "A2"

        # ── Style: Image Summary ──────────────────────────────────────────────
        ws_sum = wb["Image Summary"]
        _style_header_row(ws_sum, fill_hex=_SUMMARY_HEADER_FILL)
        _autofit_columns(ws_sum)
        ws_sum.freeze_panes = "A2"

        # Highlight min_cd and max_cd cells (Image Summary is small — per-cell ok)
        min_cd_col = _col_letter(df_summary, "min_cd_nm")
        max_cd_col = _col_letter(df_summary, "max_cd_nm")
        min_x_col  = _col_letter(df_summary, "min_cd_x_px")
        min_y_col  = _col_letter(df_summary, "min_cd_y_px")
        max_x_col  = _col_letter(df_summary, "max_cd_x_px")
        max_y_col  = _col_letter(df_summary, "max_cd_y_px")

        for row_idx in range(2, len(df_summary) + 2):
            for col_letter in (min_cd_col, min_x_col, min_y_col):
                if col_letter:
                    ws_sum[f"{col_letter}{row_idx}"].fill = min_fill
            for col_letter in (max_cd_col, max_x_col, max_y_col):
                if col_letter:
                    ws_sum[f"{col_letter}{row_idx}"].fill = max_fill

        # ── Style: Statistics ─────────────────────────────────────────────────
        ws_stat = wb["Statistics"]
        _style_header_row(ws_stat, fill_hex=_STAT_HEADER_FILL)
        _autofit_columns(ws_stat)
        ws_stat.freeze_panes = "A2"


# ── helper: filter All Measurements by export mode ───────────────────────────

def _filter_meas_by_mode(df, mode: str):
    """Return subset of All Measurements rows based on export mode.

    "all"           — unchanged (all rows)
    "min_per_image" — one row per (dataset, image_file) with lowest cd_nm
    "max_per_image" — one row per (dataset, image_file) with highest cd_nm
    """
    if mode not in ("min_per_image", "max_per_image"):
        return df
    has_dataset = (
        "dataset" in df.columns
        and not df["dataset"].astype(str).str.strip().eq("").all()
    )
    group_keys = ["dataset", "image_file"] if has_dataset else ["image_file"]
    valid = df[df["status"] != "rejected"] if "status" in df.columns else df
    if mode == "min_per_image":
        idx = valid.groupby(group_keys)["cd_nm"].idxmin()
    else:
        idx = valid.groupby(group_keys)["cd_nm"].idxmax()
    return df.loc[idx.dropna().values].copy()


# ── helper: build per-image summary ──────────────────────────────────────────

def _build_image_summary(df) -> list[dict]:
    """One row per (dataset, image_file) with stats + MIN/MAX CD locations."""
    import math

    group_keys = ["image_file"]
    if "dataset" in df.columns and not df["dataset"].astype(str).str.strip().eq("").all():
        group_keys = ["dataset", "image_file"]

    rows = []
    for keys, grp in df.groupby(group_keys, sort=False):
        if isinstance(keys, str):
            keys = (keys,)

        # Only consider non-rejected measurements
        valid = grp[grp["status"] != "rejected"]
        cd_vals = valid["cd_nm"].dropna()
        if cd_vals.empty:
            continue

        n      = len(cd_vals)
        mean   = float(cd_vals.mean())
        std    = float(cd_vals.std()) if n > 1 else 0.0
        median = float(cd_vals.median())
        mn     = float(cd_vals.min())
        mx     = float(cd_vals.max())

        # Find the row for MIN and MAX
        min_row = valid.loc[valid["cd_nm"].idxmin()]
        max_row = valid.loc[valid["cd_nm"].idxmax()]

        row: dict = {}
        if len(group_keys) > 1:
            row["dataset"] = keys[0]
        row["image_file"]    = keys[-1]
        row["nm_per_pixel"]  = float(grp["nm_per_pixel"].iloc[0]) if "nm_per_pixel" in grp.columns else 1.0
        row["recipe_name"]   = ", ".join(valid["recipe_name"].dropna().unique()) if "recipe_name" in valid.columns else ""
        row["n_measurements"] = n
        row["mean_cd_nm"]    = round(mean, 3)
        row["median_cd_nm"]  = round(median, 3)
        row["std_nm"]        = round(std, 3)
        row["3sigma_nm"]     = round(std * 3, 3)
        # MIN CD — value + location
        row["min_cd_nm"]     = round(mn, 3)
        row["min_cd_cmg_id"] = int(min_row.get("cmg_id", -1)) if "cmg_id" in min_row.index else None
        row["min_cd_col_id"] = int(min_row.get("col_id", -1)) if "col_id" in min_row.index else None
        row["min_cd_x_px"]   = float(min_row.get("cd_line_x_px", float("nan"))) if "cd_line_x_px" in min_row.index else None
        row["min_cd_y_px"]   = float(min_row.get("cd_line_y_px", float("nan"))) if "cd_line_y_px" in min_row.index else None
        # MAX CD — value + location
        row["max_cd_nm"]     = round(mx, 3)
        row["max_cd_cmg_id"] = int(max_row.get("cmg_id", -1)) if "cmg_id" in max_row.index else None
        row["max_cd_col_id"] = int(max_row.get("col_id", -1)) if "col_id" in max_row.index else None
        row["max_cd_x_px"]   = float(max_row.get("cd_line_x_px", float("nan"))) if "cd_line_x_px" in max_row.index else None
        row["max_cd_y_px"]   = float(max_row.get("cd_line_y_px", float("nan"))) if "cd_line_y_px" in max_row.index else None

        rows.append(row)
    return rows


# ── helper: build statistics table ───────────────────────────────────────────

def _build_statistics(df) -> list[dict]:
    """One section per recipe_name; falls back to one overall section."""
    import math
    rows: list[dict] = []

    recipe_col = "recipe_name" if "recipe_name" in df.columns else None
    recipes = df[recipe_col].dropna().unique().tolist() if recipe_col else ["(all)"]

    for recipe in recipes:
        if recipe_col and recipe != "(all)":
            sub = df[df[recipe_col] == recipe]
        else:
            sub = df

        valid = sub[sub["status"] != "rejected"] if "status" in sub.columns else sub
        cd_vals = valid["cd_nm"].dropna() if "cd_nm" in valid.columns else valid["y_cd_nm"].dropna()

        n    = len(cd_vals)
        mean = float(cd_vals.mean()) if n else float("nan")
        std  = float(cd_vals.std())  if n > 1 else 0.0
        rows.append({
            "recipe_name":  recipe,
            "n":            n,
            "mean_nm":      round(mean, 3) if not math.isnan(mean) else "N/A",
            "median_nm":    round(float(cd_vals.median()), 3) if n else "N/A",
            "q25_nm":       round(float(cd_vals.quantile(0.25)), 3) if n >= 2 else "N/A",
            "q75_nm":       round(float(cd_vals.quantile(0.75)), 3) if n >= 2 else "N/A",
            "std_nm":       round(std, 3),
            "3sigma_nm":    round(std * 3, 3),
            "min_nm":       round(float(cd_vals.min()), 3) if n else "N/A",
            "max_nm":       round(float(cd_vals.max()), 3) if n else "N/A",
            "n_images":     sub["image_file"].nunique() if "image_file" in sub.columns else "N/A",
        })
    return rows


# ── openpyxl style helpers ────────────────────────────────────────────────────

def _style_header_row(ws, fill_hex: str = "F7E0C8") -> None:
    from openpyxl.styles import PatternFill, Font, Alignment
    hdr_fill = PatternFill("solid", fgColor=fill_hex)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = bold
        cell.alignment = center
    ws.row_dimensions[1].height = 30


def _autofit_columns(ws, max_width: int = 30, sample_rows: int = 200) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        # Sample header + first N data rows to avoid O(n) scan over 20k+ rows
        cells_to_check = list(col_cells)[: sample_rows + 1]
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in cells_to_check
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), max_width)


def _col_index(df, col: str) -> int | None:
    """1-based column index for df column name (None if not found)."""
    try:
        return list(df.columns).index(col) + 1
    except ValueError:
        return None


def _col_letter(df, col: str) -> str | None:
    """Excel column letter for df column name (None if not found)."""
    from openpyxl.utils import get_column_letter
    idx = _col_index(df, col)
    return get_column_letter(idx) if idx is not None else None
