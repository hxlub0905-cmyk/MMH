"""Export logic for Combine Sample Measurement tool.

Three output types:
  KLARF  — new sequential DIDs, corrected XREL/YREL, original defect fields preserved
  Excel  — enriched single-sheet workbook with all columns
  Overlay — per-image PNG: ORIG crosshair (blue) → NEW crosshair (orange) + CD label
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Callable

import cv2
import numpy as np
import pandas as pd

_HERE = Path(__file__).parent
_PROJECT_ROOT = _HERE.parent.parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from src.core.klarf_writer import KlarfWriter
from tools.combine_sample_measurement.core.data_loader import build_klarf_lookup


# ── KLARF export ──────────────────────────────────────────────────────────────

def export_klarf(
    df: pd.DataFrame,
    template_parsed: dict[str, Any],
    ds_klafs: dict[str, dict[str, Any]],   # dataset_name → parsed_klarf
    output_path: str | Path,
) -> int:
    """Write a new KLARF from sampled df rows.

    Template provides header/structure; each defect row is reconstructed
    from the original KLARF entry (preserving image blocks etc.) with
    DEFECTID, XREL, YREL overwritten.

    Returns: number of defects written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    ds_lookups: dict[str, dict] = {
        name: build_klarf_lookup(pk) for name, pk in ds_klafs.items()
    }

    output_defects: list[dict[str, Any]] = []
    columns = template_parsed.get("defect_columns", [])

    for new_did, (_, row) in enumerate(df.iterrows(), start=1):
        ds_name = str(row.get("source_dataset", ""))
        stem    = Path(str(row.get("image_file", ""))).stem.lower()

        lookup  = ds_lookups.get(ds_name, {})
        entry   = lookup.get(stem)

        if entry:
            d = dict(entry["full_defect"])
        else:
            d = {col: "0" for col in columns}

        _set_field(d, "DEFECTID", "defectid", str(new_did))

        new_xrel = row.get("new_xrel")
        new_yrel = row.get("new_yrel")
        if _is_valid(new_xrel):
            _set_field(d, "XREL", "xrel", str(int(round(float(new_xrel)))))
        if _is_valid(new_yrel):
            _set_field(d, "YREL", "yrel", str(int(round(float(new_yrel)))))

        output_defects.append(d)

    KlarfWriter().write(template_parsed, output_defects, output_path)
    return len(output_defects)


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, output_path: str | Path) -> None:
    """Write enriched Excel with all columns in a single sheet."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    _COL_ORDER = [
        "new_did", "source_dataset", "image_file",
        "cd_nm", "cd_px", "axis", "nm_per_pixel",
        "cd_line_x_px", "cd_line_y_px",
        "laplacian_score", "flag",
        "old_did", "orig_xrel", "orig_yrel", "new_xrel", "new_yrel",
    ]
    out_cols = [c for c in _COL_ORDER if c in df.columns]
    out_df   = df[out_cols].reset_index(drop=True)

    _FLOAT_COLS = {
        "cd_nm", "cd_px", "laplacian_score",
        "orig_xrel", "orig_yrel", "new_xrel", "new_yrel",
    }
    for col in _FLOAT_COLS:
        if col in out_df.columns:
            out_df[col] = pd.to_numeric(out_df[col], errors="coerce").round(2)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Combined Measurements", index=False)

        ws = writer.sheets["Combined Measurements"]
        hdr_fill = PatternFill("solid", fgColor="F7E0C8")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w + 2, 32)


# ── Overlay export ────────────────────────────────────────────────────────────

def export_overlay(
    df: pd.DataFrame,
    output_dir: str | Path,
    progress_cb: Callable[[int, int], None] | None = None,
) -> list[str]:
    """Write annotated PNG for each sampled row.

    Draws:
      • Blue  crosshair + "ORIG" label  at original KLARF coordinate
      • Orange crosshair + "NEW #DID"   at corrected coordinate
      • CD value (nm) below the NEW label
      • Yellow arrow ORIG → NEW

    Returns list of written file paths.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    saved: list[str] = []
    total = len(df)

    for idx, (_, row) in enumerate(df.iterrows()):
        if progress_cb:
            progress_cb(idx, total)

        image_path = str(row.get("image_path", ""))
        if not image_path:
            continue

        img = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
        if img is None:
            continue

        # Normalise to uint8 BGR
        if img.dtype != np.uint8:
            img = cv2.normalize(img, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
        if img.ndim == 2:
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        elif img.shape[2] == 4:
            img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

        H, W = img.shape[:2]
        nm_px     = float(row.get("nm_per_pixel", 0) or 0)
        orig_xrel = row.get("orig_xrel", float("nan"))
        orig_yrel = row.get("orig_yrel", float("nan"))
        new_xrel  = row.get("new_xrel",  float("nan"))
        new_yrel  = row.get("new_yrel",  float("nan"))
        cd_nm     = float(row.get("cd_nm", 0) or 0)
        new_did   = row.get("new_did", "")

        if nm_px <= 0 or not _is_valid(orig_xrel) or not _is_valid(new_xrel):
            continue

        dx_px = (float(new_xrel) - float(orig_xrel)) / nm_px
        dy_px = (float(orig_yrel) - float(new_yrel)) / nm_px

        cx, cy  = W / 2.0, H / 2.0
        orig_pt = (int(round(cx)), int(round(cy)))
        new_pt  = (
            max(0, min(W - 1, int(round(cx + dx_px)))),
            max(0, min(H - 1, int(round(cy + dy_px)))),
        )

        annotated = img.copy()
        arm       = max(60, min(W, H) // 30)
        thickness = max(3,  min(W, H) // 400)
        font      = cv2.FONT_HERSHEY_SIMPLEX
        fs        = max(0.9, min(W, H) / 1200)
        ft        = max(2, thickness - 1)

        ORIG_COLOR  = (60,  80,  255)   # red   (BGR)
        NEW_COLOR   = (255, 200,  30)   # orange-blue (BGR → appears orange)
        ARROW_COLOR = (0,  220, 255)    # yellow

        if abs(dx_px) > 1 or abs(dy_px) > 1:
            cv2.arrowedLine(annotated, orig_pt, new_pt,
                            (0, 0, 0), thickness + 2, cv2.LINE_AA, tipLength=0.06)
            cv2.arrowedLine(annotated, orig_pt, new_pt,
                            ARROW_COLOR, thickness, cv2.LINE_AA, tipLength=0.06)
            mid = ((orig_pt[0] + new_pt[0]) // 2, (orig_pt[1] + new_pt[1]) // 2)
            dist = (dx_px ** 2 + dy_px ** 2) ** 0.5 * nm_px
            _draw_text_with_box(annotated, f"{dist:.0f} nm",
                                (mid[0] + 10, mid[1] - 10),
                                ARROW_COLOR, font, fs * 0.85, ft)

        _draw_target_marker(annotated, orig_pt,  ORIG_COLOR, arm, thickness)
        _draw_target_marker(annotated, new_pt,   NEW_COLOR,  arm, thickness)

        _draw_text_with_box(annotated, "ORIG",
                            (orig_pt[0] + int(arm * 0.7), orig_pt[1] - int(arm * 0.5)),
                            ORIG_COLOR, font, fs, ft, bold=True)

        _draw_text_with_box(annotated, f"NEW #{new_did}",
                            (new_pt[0] + int(arm * 0.7), new_pt[1] + int(arm * 0.9)),
                            NEW_COLOR, font, fs, ft, bold=True)

        _draw_text_with_box(annotated, f"{cd_nm:.2f} nm",
                            (new_pt[0] + int(arm * 0.7), new_pt[1] + int(arm * 1.85)),
                            NEW_COLOR, font, fs * 0.85, ft)

        stem     = Path(image_path).stem
        out_path = output_dir / f"{stem}_overlay.png"
        cv2.imwrite(str(out_path), annotated)
        saved.append(str(out_path))

    if progress_cb:
        progress_cb(total, total)

    return saved


# ── Drawing helpers (shared with GUI preview) ─────────────────────────────────

def draw_overlay_on_image(
    img: np.ndarray,
    nm_px: float,
    orig_xrel: float,
    orig_yrel: float,
    new_xrel: float,
    new_yrel: float,
    cd_nm: float = 0.0,
    new_did: Any = "",
) -> np.ndarray:
    """Return a copy of img with ORIG/NEW markers drawn (for GUI preview)."""
    if img.dtype != np.uint8:
        img = cv2.normalize(img, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    if img.ndim == 2:
        img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    elif img.shape[2] == 4:
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

    H, W = img.shape[:2]
    dx_px = (new_xrel - orig_xrel) / nm_px
    dy_px = (orig_yrel - new_yrel) / nm_px

    cx, cy  = W / 2.0, H / 2.0
    orig_pt = (int(round(cx)), int(round(cy)))
    new_pt  = (
        max(0, min(W - 1, int(round(cx + dx_px)))),
        max(0, min(H - 1, int(round(cy + dy_px)))),
    )

    arm       = max(60, min(W, H) // 30)
    thickness = max(3,  min(W, H) // 400)
    font      = cv2.FONT_HERSHEY_SIMPLEX
    fs        = max(0.9, min(W, H) / 1200)
    ft        = max(2, thickness - 1)

    ORIG_COLOR  = (60,  80,  255)
    NEW_COLOR   = (255, 200,  30)
    ARROW_COLOR = (0,  220, 255)

    annotated = img.copy()

    if abs(dx_px) > 1 or abs(dy_px) > 1:
        cv2.arrowedLine(annotated, orig_pt, new_pt,
                        (0, 0, 0), thickness + 2, cv2.LINE_AA, tipLength=0.06)
        cv2.arrowedLine(annotated, orig_pt, new_pt,
                        ARROW_COLOR, thickness, cv2.LINE_AA, tipLength=0.06)
        mid  = ((orig_pt[0] + new_pt[0]) // 2, (orig_pt[1] + new_pt[1]) // 2)
        dist = (dx_px ** 2 + dy_px ** 2) ** 0.5 * nm_px
        _draw_text_with_box(annotated, f"{dist:.0f} nm",
                            (mid[0] + 10, mid[1] - 10),
                            ARROW_COLOR, font, fs * 0.85, ft)

    _draw_target_marker(annotated, orig_pt,  ORIG_COLOR, arm, thickness)
    _draw_target_marker(annotated, new_pt,   NEW_COLOR,  arm, thickness)

    _draw_text_with_box(annotated, "ORIG",
                        (orig_pt[0] + int(arm * 0.7), orig_pt[1] - int(arm * 0.5)),
                        ORIG_COLOR, font, fs, ft, bold=True)
    _draw_text_with_box(annotated, f"NEW #{new_did}",
                        (new_pt[0] + int(arm * 0.7), new_pt[1] + int(arm * 0.9)),
                        NEW_COLOR, font, fs, ft, bold=True)
    _draw_text_with_box(annotated, f"{cd_nm:.2f} nm",
                        (new_pt[0] + int(arm * 0.7), new_pt[1] + int(arm * 1.85)),
                        NEW_COLOR, font, fs * 0.85, ft)

    return annotated


def _draw_target_marker(img, center, color, arm=60, thickness=3):
    x, y    = center
    outer_r = arm // 2
    inner_r = max(arm // 5, thickness * 2)

    # Outer ring — thin black outline + thin colored ring (no crosshair arms)
    ring_thick    = max(1, thickness // 2)
    outline_thick = ring_thick + 1
    cv2.circle(img, (x, y), outer_r, (0, 0, 0), outline_thick, cv2.LINE_AA)
    cv2.circle(img, (x, y), outer_r, color,      ring_thick,    cv2.LINE_AA)

    # Center filled dot
    cv2.circle(img, (x, y), inner_r + 1, (0, 0, 0), -1, cv2.LINE_AA)
    cv2.circle(img, (x, y), inner_r,     color,      -1, cv2.LINE_AA)


def _draw_text_with_box(img, text, pos, fg, font=0, font_scale=0.8,
                         thickness=2, bg=(0, 0, 0), pad=6, bold=False):
    H, W = img.shape[:2]
    if bold:
        thickness += 1
    (tw, th), baseline = cv2.getTextSize(text, font, font_scale, thickness)
    x, y = pos
    x = max(pad, min(W - tw - pad, x))
    y = max(th + pad, min(H - pad, y))
    x0, y0 = x - pad,      y - th - pad + 2
    x1, y1 = x + tw + pad, y + baseline + pad - 2
    cv2.rectangle(img, (x0, y0), (x1, y1), bg, -1,              cv2.LINE_AA)
    cv2.rectangle(img, (x0, y0), (x1, y1), fg, max(1, thickness // 2), cv2.LINE_AA)
    cv2.putText(img, text, (x, y), font, font_scale, fg, thickness, cv2.LINE_AA)


def bgr_to_pixmap(img: np.ndarray):
    from PyQt6.QtGui import QImage, QPixmap
    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    h, w, ch = rgb.shape
    qi = QImage(rgb.tobytes(), w, h, w * ch, QImage.Format.Format_RGB888)
    return QPixmap.fromImage(qi)


# ── Misc helpers ──────────────────────────────────────────────────────────────

def _is_valid(v) -> bool:
    try:
        return v is not None and not np.isnan(float(v))
    except (TypeError, ValueError):
        return False


def _set_field(d: dict, upper: str, lower: str, value: str) -> None:
    key = next((k for k in d if k.lower() == lower), None)
    if key:
        d[key] = value
    else:
        d[upper] = value
